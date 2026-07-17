#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Transcrição dos resultados de simulação (MFH_Results_v3.xlsx, iiiUC_Results.xlsx)
para os CSV normalizados do motor e4:  blocoA_resultados_mensais.csv (RB6),
blocoB_resultados_mensais.csv (iiiUC) e a atualização de metadados em *_runs.csv.

Autor: Opus 4.8 (para execução local pelo Sérgio; validar depois com dinamica.carregar()).
Data: 2026-07-16.

DECISÕES DE MAPEAMENTO (auditáveis; ver notas no fim):
  nic_util_kwh        <- Heating thermal needs
  nvc_util_kwh        <- Cooling thermal needs
  aqs_util_kwh        <- DHW thermal needs (iiiUC: 0, sem AQS)
  cons_aquecimento    <- Heating electricity           (vetor: electricidade)
  cons_arrefecimento  <- Cooling electricity           (vetor: electricidade)
  cons_iluminacao     <- Lighting electricity
  cons_equipamentos   <- Equipment + Fans + Pumps  [fans/pumps DOBRADOS aqui,
                          para que a soma das colunas eléctricas = TOTAL ELECTRICITY]
  cons_aqs / vetor_aqs / gas_kwh  -> ver INTERRUPTOR AQS_GAS_EM_CONS_AQS
  pv_producao         <- PV generation
  pv_autoconsumo      <- Generation to building
  pv_exportacao       <- |Surplus to utility|
  bateria_carga/descarga  -> 0 (a fonte só dá o net storage/perdas, não o throughput)
  rede_importacao     <- Electricity from utility (comPV) | TOTAL ELECTRICITY (semPV)
  gas_kwh             <- TOTAL GAS

Só a linha ANO é emitida: a fonte não tem perfil MENSAL de PV/necessidades/AQS,
apenas dos 6 usos finais eléctricos; o validador exige "12 meses completos ou nenhum",
pelo que se emite ANO (anual) e se omitem as linhas 01..12.  [CONFIRMAR no motor.]
"""

import csv
from openpyxl import load_workbook

# ----------------------------------------------------------------------------
# INTERRUPTOR DE CONTRATO — CONFIRMADO contra dinamica.energia_final_por_vetor():
# o motor calcula o gás a partir de cons_aqs_kwh encaminhado por vetor_aqs
# (NÃO lê a coluna gas_kwh). Logo a AQS a gás (BAU) TEM de ir em cons_aqs_kwh.
AQS_GAS_EM_CONS_AQS = True
# ----------------------------------------------------------------------------

CLIMAS = [("PRESENT DAY", "ATU"), ("2050 SSP2-4.5", "245-2050"),
          ("2050 SSP3-7.0", "370-2050"), ("2080 SSP3-7.0", "370-2080")]

BLOCOS = {
    "A": {"xlsx": "MFH_Results_v3.xlsx", "prefixo": "RB6",
          "sistemas": [("BAU Results", "BAU"), ("Multisplit Results", "MS"),
                       ("HP Results", "BC")], "tem_aqs": True},
    "B": {"xlsx": "iiiUC_Results.xlsx", "prefixo": "IIIUC",
          "sistemas": [("BAU Results", "BAU"), ("Multisplit Results", "P1"),
                       ("HP Results", "P2")], "tem_aqs": False},
}

def n(x):
    return 0.0 if x is None else float(x)

def ler_bloco(wb, sheet, clima_tag):
    rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    starts = [i for i, r in enumerate(rows) for c in r
              if isinstance(c, str) and c.strip() in [t for t, _ in CLIMAS]]
    tags = [c.strip() for r in rows for c in r
            if isinstance(c, str) and c.strip() in [t for t, _ in CLIMAS]]
    d = dict(zip(tags, starts)); s = d[clima_tag]
    e = min([x for x in sorted(d.values()) if x > s] + [len(rows)])
    out = {}
    for r in rows[s:e]:
        if len(r) > 2 and isinstance(r[1], str):
            out[r[1].strip()] = r[2]
    return out

def linha_ano(run_id, v, tem_aqs, sistema, com_pv):
    heating_el = n(v.get("Heating electricity (kWh)"))
    cooling_el = n(v.get("Cooling electricity (kWh)"))
    lighting   = n(v.get("Lighting electricity (kWh)"))
    equip      = n(v.get("Equipment electricity (kWh)"))
    fans       = n(v.get("Fans (kWh)"))
    pumps      = n(v.get("Pumps (kWh)"))
    equip_tot  = equip + fans + pumps
    total_el   = n(v.get("TOTAL ELECTRICITY (kWh)"))
    total_gas  = n(v.get("TOTAL GAS (kWh)"))
    dhw_el     = n(v.get("DHW electricity HP (kWh)")) + n(v.get("DHW electricity backup (kWh)"))
    dhw_gas    = n(v.get("DHW nat. gas (kWh)"))

    # AQS: eléctrica (MS/HP) vs gás (BAU) vs inexistente (iiiUC)
    if not tem_aqs:
        # iiiUC não tem AQS. O motor faz E[vetor_aqs]+=cons_aqs; 'none' não é
        # chave de E (KeyError) -> usa-se electricidade com 0 (inócuo).
        vetor_aqs, cons_aqs, gas = "electricidade", 0.0, 0.0
    elif sistema == "BAU":
        vetor_aqs = "gas"
        cons_aqs = dhw_gas if AQS_GAS_EM_CONS_AQS else 0.0
        gas = total_gas
    else:  # MS / HP / P1 / P2 -> AQS por bomba de calor
        vetor_aqs, cons_aqs, gas = "electricidade", dhw_el, 0.0

    if com_pv:
        pv_prod = n(v.get("PV generation (kWh)"))
        pv_auto = n(v.get("Generation to building  (kWh)"))
        pv_exp  = abs(n(v.get("Surplus to utility (kWh)")))
        rede    = n(v.get("Electricity from utility (kWh)"))
    else:
        pv_prod = pv_auto = pv_exp = 0.0
        rede = total_el  # regra do Marco para o caso sem PV

    return [run_id, "ANO",
            round(n(v.get("Heating thermal needs (kWh)")), 2),
            round(n(v.get("Cooling thermal needs (kWh)")), 2),
            round(n(v.get("DHW thermal needs (kWh)")), 2) if tem_aqs else 0.0,
            round(heating_el, 2), "electricidade",
            round(cooling_el, 2),
            round(cons_aqs, 2), vetor_aqs,
            round(lighting, 2), round(equip_tot, 2),
            round(pv_prod, 2), round(pv_auto, 2), round(pv_exp, 2),
            0.0, 0.0, round(rede, 2), round(gas, 2)]

CAB = ["run_id","periodo","nic_util_kwh","nvc_util_kwh","aqs_util_kwh",
       "cons_aquecimento_kwh","vetor_aquecimento","cons_arrefecimento_kwh",
       "cons_aqs_kwh","vetor_aqs","cons_iluminacao_kwh","cons_equipamentos_kwh",
       "pv_producao_kwh","pv_autoconsumo_kwh","pv_exportacao_kwh",
       "bateria_carga_kwh","bateria_descarga_kwh","rede_importacao_kwh","gas_kwh"]

def processar(bloco_id, base_dir="."):
    b = BLOCOS[bloco_id]
    wb = load_workbook(f"{base_dir}/{b['xlsx']}", data_only=True)
    linhas, chk = [], []
    for sheet, sysc in b["sistemas"]:
        for clima_tag, climc in CLIMAS:
            v = ler_bloco(wb, sheet, clima_tag)
            for sufixo, com_pv in [("comPV", True), ("semPV", False)]:
                clima_id = climc if climc != "ATU" else "ATU"
                run_id = f"{b['prefixo']}-{clima_id}-{sysc}-{sufixo}"
                L = linha_ano(run_id, v, b["tem_aqs"], sysc, com_pv)
                linhas.append(L)
                # self-check: soma das colunas eléctricas == TOTAL ELECTRICITY
                el = L[5] + L[7] + L[10] + L[11] + (L[8] if L[9] == "electricidade" else 0.0)
                total_el = n(v.get("TOTAL ELECTRICITY (kWh)"))
                chk.append((run_id, el, round(total_el, 2)))
    out = f"{base_dir}/bloco{bloco_id}_resultados_mensais.csv"
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f); w.writerow(CAB); w.writerows(linhas)
    return out, chk

if __name__ == "__main__":
    for bloco in ["A", "B"]:
        out, chk = processar(bloco, base_dir=".")
        print(f"\n=== Bloco {bloco} -> {out}  ({len(chk)} runs) ===")
        falhas = [(r, e, t) for r, e, t in chk if abs(e - t) > 0.5]
        for r, e, t in chk:
            flag = "" if abs(e - t) <= 0.5 else "  <-- DISCREPANCIA"
            print(f"   {r:28s} Σelec={e:9.2f}  TOTAL={t:9.2f}{flag}")
        print("   -> OK: colunas eléctricas somam o TOTAL ELECTRICITY em todos os runs."
              if not falhas else f"   -> {len(falhas)} DISCREPANCIAS!")
